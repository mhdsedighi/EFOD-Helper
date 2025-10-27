import os ,shutil ,re ,logging ,webbrowser
import win32com.client as win32
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
import xml.etree.ElementTree as ET

# Custom logging handler to output to a Tkinter Text widget
class TextHandler(logging.Handler):
    def __init__(self, text_widget, root):
        super().__init__()
        self.text_widget = text_widget
        self.root = root  # Reference to Tkinter root for updating

    def emit(self, record):
        msg = self.format(record)
        self.text_widget.configure(state='normal')
        self.text_widget.insert(tk.END, msg + '\n')
        self.text_widget.configure(state='disabled')
        self.text_widget.see(tk.END)  # Auto-scroll to the bottom
        self.root.update()  # Force GUI update to show message immediately

# Custom formatter to exclude level name except for ERROR
class CustomFormatter(logging.Formatter):
    def __init__(self):
        super().__init__(datefmt='%Y-%m-%d %H:%M:%S')

    def format(self, record):
        timestamp = self.formatTime(record, self.datefmt)
        if record.levelno == logging.ERROR:
            # Include level name for ERROR
            return f"{timestamp} - {record.levelname} - {record.msg}"
        else:
            # Exclude level name for other levels
            return f"{timestamp} - {record.msg}"

def setup_logging(text_widget, root):
    # Set up logging to output to the text widget
    logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(message)s')
    handler = TextHandler(text_widget, root)
    handler.setFormatter(CustomFormatter())
    logging.getLogger().handlers = []  # Clear default handlers
    logging.getLogger().addHandler(handler)

class Tooltip: # Tooltip class for hover text
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tooltip_window = None
        self.widget.bind("<Enter>", self.show_tooltip)
        self.widget.bind("<Leave>", self.hide_tooltip)

    def show_tooltip(self, event):
        x, y, _, _ = self.widget.bbox("insert")
        x += self.widget.winfo_rootx() + 25
        y += self.widget.winfo_rooty() + 25

        self.tooltip_window = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)  # Remove window decorations
        tw.wm_geometry(f"+{x}+{y}")

        label = tk.Label(tw, text=self.text, background="#ffffe0", relief="solid", borderwidth=1)
        label.pack()

    def hide_tooltip(self, event):
        if self.tooltip_window:
            self.tooltip_window.destroy()
            self.tooltip_window = None

def export_table_to_excel(file_path, output_dir, root):
    if not os.path.exists(file_path):
        logging.error(f"File not found: {file_path}")
        return None

    # Initialize Word application
    try:
        word = win32.Dispatch('Word.Application')
        word.Visible = False
        word.DisplayAlerts = False
        logging.info("Word application initialized")
        root.update()  # Update GUI
    except Exception as e:
        logging.error(f"Failed to initialize Word: {e}")
        return None

    try:
        # Open the document
        doc = word.Documents.Open(os.path.abspath(file_path))
        logging.info(f"Opened document: {file_path}")
        root.update()

        # Handle protection (Type 2 - Comments, no password)
        WD_NO_PROTECTION = -1
        WD_COMMENTS = 2

        if doc.ProtectionType == WD_COMMENTS:
            logging.info("Document is protected (Comments mode). Unprotecting...")
            try:
                doc.Unprotect()  # No password needed
                logging.info("Document unprotected successfully.")
                root.update()
            except Exception as e:
                logging.error(f"Failed to unprotect document: {e}")
                doc.Close()
                return None

        if doc.Tables.Count == 0:
            logging.error("No tables found in the document.")
            doc.Close()
            return None

        # Get the first table
        table = doc.Tables(1)

        # Verify Word table has 11 columns
        if table.Columns.Count != 11:
            logging.error(f"Expected 11 columns in Word table, found {table.Columns.Count}")
            doc.Close()
            return None

        # Prepare data structure
        table_data = []
        checkbox_columns = range(4, 10)  # Columns 4-9 (1-based: 4, 5, 6, 7, 8, 9)

        # Fixed checkbox text mapping
        checkbox_text_map = {
            '4': 'No Difference',
            '5': 'More Exacting',
            '6': 'Different in character',
            '7': 'Less protective or partially',
            '8': 'Significant Difference',
            '9': 'Not Applicable'
        }

        # Iterate through rows (1-based indexing)
        max_rows = table.Rows.Count
        # max_rows = min(30, table.Rows.Count)  # Limit to first 30 rows; comment out to process all rows
        logging.info(f"Processing up to {max_rows} rows")
        root.update()
        for row_idx in range(1, max_rows + 1):  # Process up to max_rows
            row_data = []
            checked_indices = []

            # Iterate through all 11 columns
            for col_idx in range(1, 12):  # 1-based: 1 to 11
                cell = table.Cell(row_idx, col_idx)
                # Get raw text
                raw_text = cell.Range.Text
                # Include tabs, newlines, and carriage returns, exclude other control characters
                visible_text = ''
                for char in raw_text:
                    if ord(char) < 32 and char not in ['\n', '\t', '\r']:
                        break
                    visible_text += char
                # Preserve all spaces, tabs, newlines, and carriage returns for columns 2, 3, 10, 11
                if col_idx in [2, 3, 10, 11]:
                    cell_text = visible_text  # No stripping to keep spaces, tabs, newlines, and carriage returns
                else:
                    cell_text = visible_text.strip()  # Strip for column 1 and others
                # Log raw and processed text for columns 2, 3, 10, 11
                if col_idx in [2, 3, 10, 11]:
                    logging.debug(f"Row {row_idx}, Col {col_idx} raw: {repr(raw_text)}, visible: {repr(visible_text)}, final: {repr(cell_text)}")
                # If in first column, extract numeric part
                if col_idx == 1:
                    numeric_match = re.match(r'^\d+(?:\.\d+)?(?![.\d])', cell_text)
                    if numeric_match:
                        cell_text = numeric_match.group(0)
                        logging.debug(
                            f"Row {row_idx}, Col {col_idx} raw: {repr(raw_text)}, visible: {repr(visible_text)}, matched: {cell_text}, remainder: {repr(raw_text[len(visible_text):].strip())}")
                    else:
                        cell_text = visible_text
                        logging.debug(
                            f"Row {row_idx}, Col {col_idx} raw: {repr(raw_text)}, visible: {repr(visible_text)}, cleaned: {cell_text} (no numeric match)")
                # Handle checkbox columns (4-9)
                if col_idx in checkbox_columns:
                    for field in cell.Range.FormFields:
                        if field.Type == 71:  # wdFieldFormCheckBox
                            if field.CheckBox.Value:
                                checked_indices.append(str(col_idx))
                            # Debug: Print raw cell content if problematic
                            if any(ord(c) < 32 and c not in ['\n', '\t', '\r'] for c in raw_text):
                                logging.debug(f"Row {row_idx}, Col {col_idx} raw content: {repr(raw_text)}")
                # Add specific columns to row_data in desired order
                elif col_idx in [1, 2, 3, 10, 11]:  # Only these go to Excel
                    if col_idx == 1:
                        row_data.append(cell_text)  # Annex Ref. (Excel col 1)
                    elif col_idx == 2:
                        row_data.append(cell_text)  # Standard (Excel col 2)
                    elif col_idx == 3:
                        row_data.append(cell_text)  # State Ref. (Excel col 4)
                    elif col_idx == 10:
                        row_data.append(cell_text)  # Details (Excel col 5)
                    elif col_idx == 11:
                        row_data.append(cell_text)  # Remark (Excel col 6)

            # Determine text for checked checkboxes
            if len(checked_indices) == 0:
                checked_text = ""  # No checkboxes checked
            elif len(checked_indices) == 1:
                checked_text = checkbox_text_map.get(checked_indices[0], "")
            else:
                checked_text = "error-multi checkbox"

            # Insert checked text as the third column (Excel col 3)
            row_data.insert(2, checked_text)  # Inserts "Difference" at index 2
            table_data.append(row_data)
            root.update()  # Update GUI after each row

        # Define exactly 6 column headers for Excel
        headers = [
            "Annex Ref.",      # Word col 1
            "Standard",        # Word col 2
            "Difference",      # Checkboxes from Word cols 4-9
            "State Ref.",      # Word col 3
            "Details",         # Word col 10
            "Remark"           # Word col 11
        ]

        # Create DataFrame with exactly 6 columns
        df = pd.DataFrame(table_data, columns=headers)

        # Generate output filename
        base_name = os.path.splitext(os.path.basename(file_path))[0] + ".xlsx"
        output_excel_path = os.path.join(output_dir, base_name)
        counter = 1
        while os.path.exists(output_excel_path):
            output_excel_path = os.path.join(output_dir, f"{os.path.splitext(base_name)[0]}_{counter}.xlsx")
            counter += 1

        # Export to Excel initially
        df.to_excel(output_excel_path, index=False, engine='openpyxl')

        # Load the workbook to modify it
        wb = load_workbook(output_excel_path)
        ws = wb.active

        # Define the table range (A1 to F<rows+1> for 6 columns)
        num_rows = len(table_data) + 1  # +1 for header
        table_range = f"A1:F{num_rows}"

        # Create an Excel table
        tab = Table(displayName="FormDataTable", ref=table_range)
        style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        ws.add_table(tab)

        # Freeze the first row (headers)
        ws.freeze_panes = ws['A2']  # Freezes row 1

        # Save the modified workbook
        wb.save(output_excel_path)
        logging.info(f"Table data exported to {output_excel_path} (Processed {len(table_data)} rows) with table and frozen headers")
        root.update()
        return output_excel_path

    except Exception as e:
        logging.error(f"An error occurred in export_table_to_excel: {e}")
        return None

    finally:
        try:
            doc.Close()
            logging.info("Document closed")
            root.update()
        except:
            pass
        try:
            word.Quit()
            logging.info("Word application quit")
            root.update()
        except:
            pass

def fill_form_from_excel(excel_path, form_path, root):
    WD_NO_PROTECTION = -1
    WD_COMMENTS = 2
    if not os.path.exists(excel_path):
        logging.error(f"Excel file not found: {excel_path}")
        return None
    if not os.path.exists(form_path):
        logging.error(f"Form file not found: {form_path}")
        return None

    # Read Excel data
    try:
        df = pd.read_excel(excel_path)
        if list(df.columns) != ["Annex Ref.", "Standard", "Difference", "State Ref.", "Details", "Remark"]:
            logging.error("Excel file does not match expected column structure")
            return None
        logging.info("Excel data loaded successfully")
        root.update()
    except Exception as e:
        logging.error(f"Failed to read Excel file: {e}")
        return None

    # Validate "Difference" column values
    checkbox_text_map = {
        'no': 4,  # Short form
        'no difference': 4,  # Full form
        'more': 5,  # Short form
        'more exacting': 5,  # Full form
        'more exacting or exceeds': 5,  # Full form
        'different': 6,  # Short form
        'different in character': 6,  # Full form
        'difference in character': 6,  # Full form
        'difference in character or Other means of compliance': 6,  # Full form
        'less': 7,  # Short form
        'less protective or partially': 7,  # short form
        'Less protective or Partially Implemented or Not Implemented': 7,  # Full form
        'significant': 8,  # Short form
        'significant difference': 8,  # Full form
        'not': 9,  # Short form
        'not applicable': 9,  # Full form
        'not defined': 9  # Full form
    }
    valid_values = list(checkbox_text_map.keys()) + ['error-multi checkbox',
                                                     '']  # Allow empty string and error-multi checkbox
    invalid_rows = []
    for idx, value in enumerate(df["Difference"]):
        # Convert value to string and normalize for comparison
        diff_value = str(value).strip().lower() if pd.notna(value) else ''
        if diff_value not in valid_values:
            invalid_rows.append((idx + 2, value))  # Excel row number (index + 2 due to header)

    if invalid_rows:
        error_message = "Invalid values found in the 'Difference' column. The following rows contain unrecognized values:\n\n"
        for row_num, value in invalid_rows:
            error_message += f"Row {row_num}: '{value}'\n"
        error_message += "\nExpected values are: " + ", ".join(
            f"'{v}'" for v in checkbox_text_map.keys()) + ", 'error-multi checkbox', or empty."
        logging.error(error_message)
        messagebox.showerror("Invalid Difference Values", error_message)
        return None

    # Create backup of the form file
    try:
        output_dir = os.path.dirname(form_path)
        base_name = os.path.splitext(os.path.basename(form_path))[0]
        backup_path = os.path.join(output_dir, f"{base_name}_beforefilling.docx")
        counter = 1
        while os.path.exists(backup_path):
            backup_path = os.path.join(output_dir, f"{base_name}_beforefilling_{counter}.docx")
            counter += 1
        shutil.copy2(form_path, backup_path)  # Copy the file preserving metadata
        logging.info(f"Created backup: {backup_path}")
        root.update()
    except Exception as e:
        logging.error(f"Failed to create backup: {e}")
        return None

    # Initialize Word application
    try:
        word = win32.Dispatch('Word.Application')
        word.Visible = False
        word.DisplayAlerts = False
        logging.info("Word application initialized")
        root.update()
    except Exception as e:
        logging.error(f"Failed to initialize Word: {e}")
        return None

    try:
        # Open the original document
        logging.debug(f"Attempting to open document: {form_path}")
        doc = word.Documents.Open(os.path.abspath(form_path))
        logging.info(f"Opened document: {form_path}")
        root.update()

        # Get the first table
        if doc.Tables.Count == 0:
            logging.error("No tables found in the document.")
            doc.Close()
            return None

        table = doc.Tables(1)

        # Verify Word table has 11 columns
        logging.debug("Verifying column count")
        if table.Columns.Count != 11:
            logging.error(f"Expected 11 columns in Word table, found {table.Columns.Count}")
            doc.Close()
            return None

        # Check if number of rows matches
        excel_rows = len(df)
        word_rows = table.Rows.Count
        if excel_rows != word_rows:
            logging.error(f"Row count mismatch: Excel has {excel_rows} rows, Word table has {word_rows} rows")
            messagebox.showerror("Error", f"Row count mismatch: Excel has {excel_rows} rows, Word table has {word_rows} rows")
            doc.Close()
            return None
        logging.info(f"Row count matches: {excel_rows} rows in both Excel and Word table")
        root.update()

        # Iterate through rows
        max_rows = table.Rows.Count
        logging.info(f"Processing {max_rows} rows")
        root.update()
        for row_idx in range(1, max_rows + 1):
            row_data = df.iloc[row_idx - 1]  # 0-based index for pandas

            # Fill text fields (columns 3, 10, 11)
            for col_idx, value in [(3, row_data["State Ref."]), (10, row_data["Details"]), (11, row_data["Remark"])]:
                try:
                    cell = table.Cell(row_idx, col_idx)
                    if cell.Range.FormFields.Count > 0:
                        field = cell.Range.FormFields(1)
                        if field.Type == 70:  # wdFieldFormTextInput
                            # Handle NaN values
                            if pd.isna(value):
                                cell_text = ""
                            else:
                                # Convert to string and clean up whitespace
                                cell_text = str(value)

                                # Remove leading/trailing whitespace (including Unicode spaces)
                                cell_text = cell_text.strip()

                                # Remove internal Unicode whitespace characters that are effectively empty
                                # Keep regular spaces, tabs, and newlines, but remove exotic Unicode spaces
                                if not cell_text or len(cell_text) <= 3:
                                    # For very short strings, check if it's just whitespace
                                    cleaned = re.sub(r'[\u2000-\u200B\u2028\u2029\u202F\u205F\u3000]+', '', cell_text)
                                    if not cleaned:
                                        cell_text = ""  # Treat as empty if only Unicode whitespace
                                    else:
                                        cell_text = cleaned  # Keep the cleaned content
                                else:
                                    # For longer content, just strip and keep as-is
                                    pass

                            # Set the field
                            field.Result = cell_text

                            if cell_text:logging.debug(f"Set text field in Row {row_idx}, Col {col_idx}: '{cell_text}'")
                            else:
                                logging.debug(f"Cleared text field in Row {row_idx}, Col {col_idx} (empty)")
                        else:
                            logging.warning(f"Expected text field, found type {field.Type} in Row {row_idx}, Col {col_idx}")
                    else:
                        logging.warning(f"No form field in Row {row_idx}, Col {col_idx} - skipping")
                except Exception as e:
                    logging.error(f"Failed to set text in Row {row_idx}, Col {col_idx}: {e}")
            root.update()

            # Handle checkboxes (columns 4-9)
            diff_value = row_data["Difference"]
            expected_col = None

            # First, determine the current checkbox state - detect ALL checked boxes
            current_checked_cols = []
            for col_idx in range(4, 10):  # Columns 4-9
                try:
                    cell = table.Cell(row_idx, col_idx)
                    if cell.Range.FormFields.Count > 0:
                        for field in cell.Range.FormFields:
                            if field.Type == 71:  # wdFieldFormCheckBox
                                if field.CheckBox.Value:
                                    current_checked_cols.append(col_idx)
                                    # logging.debug(f"Row {row_idx}: Found checked checkbox in Col {col_idx}")
                except Exception as e:
                    logging.error(f"Error reading checkbox in Row {row_idx}, Col {col_idx}: {e}")

            # Determine what the checkbox state should be
            if pd.notna(diff_value) and diff_value != "error-multi checkbox":
                diff_key = str(diff_value).strip().lower()
                if diff_key in checkbox_text_map:
                    expected_col = checkbox_text_map[diff_key]
                else:
                    # Unrecognized value - should all be unchecked
                    expected_col = None
            else:
                # NaN or "error-multi checkbox" - should all be unchecked
                expected_col = None

            # # Log current and expected state
            # if len(current_checked_cols) > 1:
            #     logging.warning(f"Row {row_idx}: Multiple checkboxes detected! Currently checked: {current_checked_cols}, Expected: {expected_col or 'none'}")
            # elif len(current_checked_cols) == 1:
            #     logging.debug(f"Row {row_idx}: Single checkbox currently checked in Col {current_checked_cols[0]}, Expected: {expected_col or 'none'}")
            # else:
            #     logging.debug(f"Row {row_idx}: No checkboxes currently checked, Expected: {expected_col or 'none'}")

            # Determine if any changes are needed
            needs_change = False
            if expected_col is None:
                # Expected: all unchecked
                needs_change = len(current_checked_cols) > 0
            else:
                # Expected: specific one checked
                needs_change = (len(current_checked_cols) != 1) or (current_checked_cols[0] != expected_col)

            if not needs_change:
                logging.debug(f"Row {row_idx}: No Checkbox changes needed")
            else:
                # logging.debug(f"Row {row_idx}: Checkbox state change needed - current: {current_checked_cols}, expected: {expected_col or 'none'}")

                # Whenever change is needed, first Uncheck ALL currently checked boxes
                for col_idx in current_checked_cols:
                    try:
                        cell = table.Cell(row_idx, col_idx)
                        if cell.Range.FormFields.Count > 0:
                            for field in cell.Range.FormFields:
                                if field.Type == 71:  # wdFieldFormCheckBox
                                    field.CheckBox.Value = False
                                    logging.debug(f"Row {row_idx}: Unchecked box in Col {col_idx}")
                    except Exception as e:
                        logging.error(f"Row {row_idx}: Error unchecking checkbox in Col {col_idx}: {e}")

                # Check the expected one (if any)
                if expected_col:
                    try:
                        cell = table.Cell(row_idx, expected_col)
                        if cell.Range.FormFields.Count > 0:
                            for field in cell.Range.FormFields:
                                if field.Type == 71:  # wdFieldFormCheckBox
                                    field.CheckBox.Value = True
                                    logging.debug(f"Row {row_idx}: Checked box in Col {expected_col}: {diff_value}")
                    except Exception as e:
                        logging.error(f"Row {row_idx}: Error setting checkbox in Col {expected_col}: {e}")

                # Comprehensive verification - check ALL checkboxes in columns 4-9
                # logging.debug(f"Row {row_idx}: Verifying final checkbox state...")
                final_checked_cols = []
                for col_idx in range(4, 10):
                    try:
                        cell = table.Cell(row_idx, col_idx)
                        if cell.Range.FormFields.Count > 0:
                            for field in cell.Range.FormFields:
                                if field.Type == 71:  # wdFieldFormCheckBox
                                    if field.CheckBox.Value:
                                        final_checked_cols.append(col_idx)
                                        if col_idx != expected_col:
                                            # Unexpected checkbox is still checked - force uncheck
                                            # logging.warning(f"Row {row_idx}: Unexpected checkbox found in Col {col_idx} - force unchecking")
                                            field.CheckBox.Value = False
                                            final_checked_cols.remove(
                                                col_idx)  # Remove from list since we just fixed it
                    except Exception as e:
                        logging.error(f"Row {row_idx}: Error verifying checkbox in Col {col_idx}: {e}")

                # Double-check the expected one is actually checked
                if expected_col:
                    try:
                        cell = table.Cell(row_idx, expected_col)
                        if cell.Range.FormFields.Count > 0:
                            for field in cell.Range.FormFields:
                                if field.Type == 71 and not field.CheckBox.Value:
                                    logging.error(
                                        f"Row {row_idx}: Expected checkbox in Col {expected_col} is not checked - force setting!")
                                    field.CheckBox.Value = True
                                    final_checked_cols.append(expected_col)
                    except Exception as e:
                        logging.error(f"Row {row_idx}: Error double-checking checkbox in Col {expected_col}: {e}")

                # Final state validation
                if expected_col is None:
                    # Should be all unchecked
                    if final_checked_cols:
                        logging.error(f"Row {row_idx}: Verification failed! Expected all unchecked but found {final_checked_cols} still checked")
                    else:
                        logging.debug(f"Row {row_idx}: Verification passed - all checkboxes unchecked as expected")
                else:
                    # Should have exactly one checked (the expected one)
                    if len(final_checked_cols) == 1 and final_checked_cols[0] == expected_col:
                        logging.debug(
                            f"Row {row_idx}: Verification passed - Col {expected_col} checked, others unchecked")
                    else:
                        logging.error(f"Row {row_idx}: Verification failed! Expected Col {expected_col} checked, but found {final_checked_cols}")

            root.update()

        # Save changes to the original document
        doc.Save()
        logging.info(f"Changes saved to original document: {form_path}")

        # Re-apply original protection (Type 2 - Comments, no password)
        if doc.ProtectionType == WD_NO_PROTECTION:  # Check if we unprotected it
            try:
                doc.Protect(Type=WD_COMMENTS, NoReset=False)
                logging.info("Document re-protected (Comments mode)")
            except Exception as e:
                logging.warning(f"Failed to re-apply protection: {e}")

        root.update()
        return form_path

    except Exception as e:
        logging.error(f"An error occurred in fill_form_from_excel: {e}")
        return None

    finally:
        try:
            doc.Close()
            logging.info("Document closed")
            root.update()
        except:
            pass
        try:
            word.Quit()
            logging.info("Word application quit")
            root.update()
        except:
            pass

def xml_to_excel(xml_path, output_dir, root):
    if not os.path.exists(xml_path):
        logging.error(f"XML file not found: {xml_path}")
        return None

    try:
        # Parse the XML file
        tree = ET.parse(xml_path)
        xml_root = tree.getroot()

        # Define namespaces
        namespaces = {'ns': 'urn:crystal-reports:schemas:report-detail'}

        # Extract relevant data
        table_data = []
        for details in xml_root.findall('.//ns:Details', namespaces):
            row_data = []
            annex_ref = details.find('.//ns:Field[@Name="AnnexReferenceNumber1"]/ns:Value', namespaces)
            standard = details.find('.//ns:Field[@Name="SARP11"]/ns:Value', namespaces)
            state_ref = details.find('.//ns:Field[@Name="StateReference1"]/ns:Value', namespaces)
            difference = details.find('.//ns:Field[@Name="StateDifferenceLevel1"]/ns:Value', namespaces)
            state_difference = details.find('.//ns:Field[@Name="StateDifference1"]/ns:Value', namespaces)
            state_comments = details.find('.//ns:Field[@Name="StateComments1"]/ns:Value', namespaces)

            # Check if elements are found and extract text
            annex_ref_text = annex_ref.text if annex_ref is not None else "Not Found"
            standard_text = standard.text if standard is not None else "Not Found"
            state_ref_text = state_ref.text if state_ref is not None else "Not Found"
            difference_text = difference.text if difference is not None else "Not Found"
            if difference_text.lower().startswith("less p"):
                difference_text = "Less protective or Partially Implemented or Not Implemented"
            elif difference_text.lower().startswith("more e"):
                difference_text = "More Exacting or Exceeds"
            elif difference_text.lower().startswith("difference"):
                difference_text = "Difference in character or Other means of compliance"
            state_difference_text = state_difference.text if state_difference is not None else "Not Found"
            state_comments_text = state_comments.text if state_comments is not None else "Not Found"

            # Log the extracted data
            logging.debug(f"Annex Ref: {annex_ref_text}, Standard: {standard_text}, State Ref: {state_ref_text}, "
                          f"Difference: {difference_text}, Details: {state_difference_text}, Remark: {state_comments_text}")

            row_data.append(annex_ref_text)
            row_data.append(standard_text)
            row_data.append(difference_text)
            row_data.append(state_ref_text)
            row_data.append(state_difference_text)
            row_data.append(state_comments_text)

            table_data.append(row_data)

        if not table_data:
            logging.error("No data extracted from XML.")
            return None

        # Define column headers for Excel
        headers = [
            "Annex Ref.",
            "Standard",
            "Difference",
            "State Ref.",
            "Details",
            "Remark"
        ]

        # Create DataFrame
        df = pd.DataFrame(table_data, columns=headers)
        logging.debug(f"DataFrame created with data:\n{df}")

        # Generate unique output filename
        base_name = "output_from_xml.xlsx"
        output_excel_path = os.path.join(output_dir, base_name)
        counter = 1
        while os.path.exists(output_excel_path):
            output_excel_path = os.path.join(output_dir, f"output_from_xml_{counter}.xlsx")
            counter += 1

        # Export to Excel initially
        df.to_excel(output_excel_path, index=False, engine='openpyxl')

        # Load the workbook to modify it
        wb = load_workbook(output_excel_path)
        ws = wb.active

        # Define the table range (A1 to F<rows+1> for 6 columns)
        num_rows = len(table_data) + 1  # +1 for header
        table_range = f"A1:F{num_rows}"

        # Create an Excel table
        tab = Table(displayName="FormDataTable", ref=table_range)
        style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        ws.add_table(tab)

        # Freeze the first row (headers)
        ws.freeze_panes = ws['A2']  # Freezes row 1

        # Save the modified workbook
        wb.save(output_excel_path)
        logging.info(f"XML data exported to {output_excel_path} (Processed {len(table_data)} rows) with table and frozen headers")
        root.update()
        return output_excel_path

    except Exception as e:
        logging.error(f"An error occurred in xml_to_excel: {e}")
        return None


def excel_on_excel(sample_excel_path, fillable_excel_path, root):
    if not os.path.exists(sample_excel_path):
        logging.error(f"Sample Excel file not found: {sample_excel_path}")
        return None
    if not os.path.exists(fillable_excel_path):
        logging.error(f"Fillable Excel file not found: {fillable_excel_path}")
        return None

    # Read sample Excel data
    try:
        sample_df = pd.read_excel(sample_excel_path, dtype={0: str})
        logging.info(f"Sample Excel loaded: {len(sample_df)} rows, columns: {list(sample_df.columns)}")
        logging.debug(f"Sample first column values: {sample_df.iloc[:, 0].tolist()}")
        root.update()
    except Exception as e:
        logging.error(f"Failed to read sample Excel: {e}")
        return None

    # Read fillable Excel data
    try:
        fillable_df = pd.read_excel(fillable_excel_path, dtype={0: str})
        logging.info(f"Fillable Excel loaded: {len(fillable_df)} rows, columns: {list(fillable_df.columns)}")
        logging.debug(f"Fillable first column values: {fillable_df.iloc[:, 0].tolist()}")
        root.update()
    except Exception as e:
        logging.error(f"Failed to read fillable Excel: {e}")
        return None

    # Create sample dictionary
    try:
        sample_dict = {str(row.iloc[0]).strip(): row for index, row in sample_df.iterrows()}
        logging.info(f"Sample dictionary created with {len(sample_dict)} unique keys")
        logging.debug(f"Sample dictionary keys: {list(sample_dict.keys())}")
    except Exception as e:
        logging.error(f"Failed to create sample dictionary: {e}")
        return None

    # Process fillable_df
    processed_rows = 0
    for index, row in fillable_df.iterrows():
        try:
            logging.debug(f"Processing row {index + 2} (index {index})")
            # Validate first column
            if pd.isna(row.iloc[0]):
                logging.warning(f"Row {index + 2}: First column is NaN, skipping update")
                continue

            fillable_ref = str(row.iloc[0]).strip()
            logging.debug(f"Row {index + 2}: fillable_ref = {repr(fillable_ref)}")

            # Check DataFrame state
            logging.debug(f"Row {index + 2}: fillable_df shape before update: {fillable_df.shape}")

            sample_row = sample_dict.get(fillable_ref)
            if sample_row is not None:
                # Log sample row data
                logging.debug(f"Row {index + 2}: Found match, sample_row = {sample_row.tolist()}")
                # Update specific columns, preserving column 2 ("Standard") from fillable_df
                fillable_df.iloc[index, 0] = sample_row.iloc[0]  # Annex Ref.
                fillable_df.iloc[index, 2] = sample_row.iloc[2]  # Difference
                fillable_df.iloc[index, 3] = sample_row.iloc[3]  # State Ref.
                fillable_df.iloc[index, 4] = sample_row.iloc[4]  # Details
                fillable_df.iloc[index, 5] = sample_row.iloc[5]  # Remark
                logging.info(f"Row {index + 2}: Updated columns 1,3,4,5,6 for fillable_ref = {fillable_ref}, preserved Standard = {repr(fillable_df.iloc[index, 1])}")
            else:
                logging.debug(f"Row {index + 2}: No match found for fillable_ref = {fillable_ref}")

            # Increment counter
            processed_rows += 1
            logging.debug(f"Row {index + 2}: fillable_df shape after update: {fillable_df.shape}")

            root.update()
        except Exception as e:
            logging.error(f"Error processing row {index + 2}: {e}")
            logging.debug(f"Row {index + 2}: row data = {row.tolist()}")
            raise  # Re-raise to stop and inspect

    logging.info(f"Processed {processed_rows} rows out of {len(fillable_df)}")

    # Generate output filename
    output_dir = os.path.dirname(fillable_excel_path)
    base_name = os.path.splitext(os.path.basename(fillable_excel_path))[0] + "_filled.xlsx"
    output_excel_path = os.path.join(output_dir, base_name)
    counter = 1
    while os.path.exists(output_excel_path):
        output_excel_path = os.path.join(output_dir, f"{os.path.splitext(os.path.basename(fillable_excel_path))[0]}_filled_{counter}.xlsx")
        counter += 1

    # Export to Excel
    try:
        fillable_df.to_excel(output_excel_path, index=False, engine='openpyxl')
        logging.debug(f"Excel saved to {output_excel_path}")

        # Load workbook to add table
        wb = load_workbook(output_excel_path)
        ws = wb.active
        num_rows = len(fillable_df) + 1
        num_cols = len(fillable_df.columns)
        table_range = f"A1:{chr(64 + num_cols)}{num_rows}"
        tab = Table(displayName="FormDataTable", ref=table_range)
        style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        ws.add_table(tab)
        ws.freeze_panes = ws['A2']
        wb.save(output_excel_path)
        logging.info(f"Excel saved: {output_excel_path} ({len(fillable_df)} rows)")
        root.update()
        return output_excel_path
    except Exception as e:
        logging.error(f"Failed to save Excel: {e}")
        return None


def gui():
    root = tk.Tk()
    root.title("EFOD Helper")
    root.state('zoomed')  # Maximize window on startup
    root.configure(bg="#1C2526")  # Dark background

    # Frame for buttons
    button_frame = tk.Frame(root, bg="#1C2526")
    button_frame.pack(pady=10, fill=tk.X)  # Fill horizontally

    # Log display area (dynamic width)
    log_text = scrolledtext.ScrolledText(root, height=20, state='disabled',bg="#2E2E2E", fg="#E0E0E0",insertbackground="#E0E0E0", font=("Arial", 10))
    log_text.pack(pady=10, fill=tk.BOTH, expand=True)  # Fill both directions and expand

    # Set up logging to the text widget
    setup_logging(log_text, root)

    def form_to_excel():
        form_path = filedialog.askopenfilename(title="Select Word Form", filetypes=[("Word files", "*.docx")])
        if form_path:
            output_dir = os.path.dirname(form_path)
            output_file = export_table_to_excel(form_path, output_dir, root)
            if output_file:
                messagebox.showinfo("Success", f"Conversion completed. Output saved as: {output_file}",parent=root)
            else:
                messagebox.showerror("Error", "Conversion failed. Check logs for details.", parent=root)

    def excel_to_form():
        excel_path = filedialog.askopenfilename(title="Select Excel File", filetypes=[("Excel files", "*.xlsx")])
        if excel_path:
            form_path = filedialog.askopenfilename(title="Select EFOD Form to Edit", filetypes=[("Word files", "*.docx")])
            if form_path:
                output_file = fill_form_from_excel(excel_path, form_path, root)
                if output_file:
                    messagebox.showinfo("Success", f"Form filled and saved as: {output_file}", parent=root)
                else:
                    messagebox.showerror("Error", "Conversion failed. Check logs for details.", parent=root)

    def xml_to_excel_conversion():
        xml_path = filedialog.askopenfilename(title="Select XML File of a country, Exported from SAP Crystal Reports", filetypes=[("XML files", "*.xml")])
        if xml_path:
            output_dir = os.path.dirname(xml_path)
            output_file = xml_to_excel(xml_path, output_dir, root)
            if output_file:
                messagebox.showinfo("Success", f"Conversion completed. Output saved as: {output_file}", parent=root)
            else:
                messagebox.showerror("Error", "Conversion failed. Check logs for details.", parent=root)

    def excel_on_excel_conversion():
        sample_excel_path = filedialog.askopenfilename(title="Select Sample Excel File (to read from)",
                                                       filetypes=[("Excel files", "*.xlsx")])
        if sample_excel_path:
            fillable_excel_path = filedialog.askopenfilename(title="Select Fillable Excel File",
                                                             filetypes=[("Excel files", "*.xlsx")])
            if fillable_excel_path:
                output_file = excel_on_excel(sample_excel_path, fillable_excel_path, root)
                if output_file:
                    messagebox.showinfo("Success", f"Excel filled and saved as: {output_file}", parent=root)
                else:
                    messagebox.showerror("Error", "Conversion failed. Check logs for details.", parent=root)

    def show_help_dialog():
        # Create a custom dialog box
        help_dialog = tk.Toplevel(root)
        help_dialog.title("Help")
        help_dialog.geometry("400x200")
        help_dialog.transient(root)
        help_dialog.grab_set()
        help_dialog.configure(bg="#1C2526")

        # Text message
        message = tk.Label(help_dialog, text="For more information and getting the latest update visit my GitHub:",
                           wraplength=350, justify="center", bg="#1C2526", fg="#E0E0E0", font=("Arial", 10))
        message.pack(pady=20)

        # Clickable link
        link_text = "https://github.com/mhdsedighi/EFOD-Helper"
        link = tk.Label(help_dialog, text=link_text, fg="#4FC3F7", cursor="hand2", font=("Arial", 10, "underline"),bg="#1C2526")
        link.pack(pady=10)
        link.bind("<Button-1>", lambda e: webbrowser.open(link_text))

        # OK button
        ok_button = tk.Button(help_dialog, text="OK", command=help_dialog.destroy, width=10,bg="#4CAF50", fg="#E0E0E0", activebackground="#66BB6A",font=("Arial", 10), bd=0)
        ok_button.pack(pady=20)

    # Configure button style
    button_style = {
        "bg": "#37474F",           # Dark gray button background
        "fg": "#E0E0E0",           # Light text
        "activebackground": "#546E7A",  # Slightly lighter on hover
        "activeforeground": "#E0E0E0",
        "font": ("Arial", 10),
        "bd": 0,                   # Borderless for modern look
        "relief": "flat",
        "padx": 10,
        "pady": 5
    }

    # Buttons
    btn_form_to_excel = tk.Button(button_frame, text="EFOD → Excel", command=form_to_excel, width=20, **button_style)
    btn_form_to_excel.pack(side=tk.LEFT, padx=10)
    Tooltip(btn_form_to_excel, "Convert an EFOD Word form to an Excel file")

    btn_excel_to_form = tk.Button(button_frame, text="Excel → EFOD", command=excel_to_form, width=20, **button_style)
    btn_excel_to_form.pack(side=tk.LEFT, padx=10)
    Tooltip(btn_excel_to_form, "Fill an EFOD Word form with the data from an Excel file")

    btn_xml_to_excel = tk.Button(button_frame, text="SAP Crystal Reports → Excel", command=xml_to_excel_conversion, width=20, **button_style)
    btn_xml_to_excel.pack(side=tk.LEFT, padx=10)
    Tooltip(btn_xml_to_excel, "Convert a SAP Crystal Reports XML export (for a country) to an Excel file")

    btn_excel_on_excel = tk.Button(button_frame, text="Excel → Excel", command=excel_on_excel_conversion, width=20, **button_style)
    btn_excel_on_excel.pack(side=tk.LEFT, padx=10)
    Tooltip(btn_excel_on_excel, "Fill one Excel file with data from another Excel file based on matching annex ref. number")

    # Help button
    btn_help = tk.Button(button_frame, text="?", command=show_help_dialog, width=5,bg="#0288D1", fg="#E0E0E0", activebackground="#03A9F4",font=("Arial", 10), bd=0, relief="flat")
    btn_help.pack(side=tk.LEFT, padx=10)
    Tooltip(btn_help, "More info")

    # Update Tooltip class to match theme
    def update_tooltip_style():
        Tooltip.show_tooltip = lambda self, event: (
            setattr(self, 'tooltip_window', tw := tk.Toplevel(self.widget)),
            tw.wm_overrideredirect(True),
            tw.wm_geometry(f"+{self.widget.winfo_rootx() + 25}+{self.widget.winfo_rooty() + 25}"),
            tk.Label(tw, text=self.text, background="#37474F", foreground="#E0E0E0",
                     font=("Arial", 9), relief="solid", borderwidth=1).pack()
        )

    update_tooltip_style()

    root.mainloop()


if __name__ == "__main__":
    gui()